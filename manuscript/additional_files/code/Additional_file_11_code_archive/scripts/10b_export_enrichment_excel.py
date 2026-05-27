#!/usr/bin/env python
"""Create a verified Excel workbook for enrichment analysis outputs."""

from pathlib import Path

import pandas as pd


PROJECT_DIR = Path(__file__).resolve().parents[1]
ENRICHMENT_DIR = PROJECT_DIR / "results" / "enrichment"
OUT = ENRICHMENT_DIR / "enrichment_summary.xlsx"

SHEETS = {
    "triple_qc": ENRICHMENT_DIR / "enrichment_triple_qc.csv",
    "database_qc": ENRICHMENT_DIR / "enrichment_database_qc.csv",
    "gene_set_inclusion_qc": ENRICHMENT_DIR / "gene_set_inclusion_qc.csv",
    "gene_sets_used": ENRICHMENT_DIR / "gene_sets_used.csv",
    "significant_enrichment": ENRICHMENT_DIR / "all_enrichment_significant_fdr0.05.csv",
    "all_enrichment": ENRICHMENT_DIR / "all_enrichment_results.csv",
}


def main() -> None:
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for sheet_name, path in SHEETS.items():
            if not path.exists():
                raise FileNotFoundError(path)
            df = pd.read_csv(path)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.style = "Headline 4"

    from openpyxl import load_workbook

    wb = load_workbook(OUT, read_only=True, data_only=True)
    for sheet_name, path in SHEETS.items():
        ws = wb[sheet_name]
        if ws.max_row < 2 and path.stat().st_size > 5:
            raise RuntimeError(f"Sheet {sheet_name} looks empty after export.")
        print(f"{sheet_name}: {ws.max_row} rows x {ws.max_column} columns")


if __name__ == "__main__":
    main()
