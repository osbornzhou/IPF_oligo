#!/usr/bin/env python
"""Merge curated annotation workbooks and check sample IDs against expression matrices.

This script discovers Excel workbooks under data_raw/GEO that contain an
`annotation_analysis` sheet, merges them, and checks whether `sample_id`
matches the sample columns of the expression file used for each dataset.
"""

from __future__ import annotations

import csv
import gzip
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
GEO_ROOT = PROJECT_ROOT / "data_raw" / "GEO"
METADATA_ROOT = PROJECT_ROOT / "metadata"
LOG_ROOT = PROJECT_ROOT / "logs"

ANNOTATION_SHEET = "annotation_analysis"
MIN_STANDARD_COLUMNS = {"sample_id", "series_id", "data_type", "group", "include", "platform"}

STANDARD_COLUMNS = [
    "sample_id",
    "geo_accession",
    "series_id",
    "data_type",
    "tissue",
    "organism",
    "age",
    "sex",
    "group",
    "subgroup",
    "include",
    "smoke",
    "platform",
    "rin",
    "fvc_percent_predicted",
    "dlco_percent_predicted",
    "repository",
    "preservative",
    "dataset_role",
    "batch",
    "notes",
]


def workbook_sheets(path: Path) -> list[str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        return wb.sheetnames
    except Exception:
        return []


def find_annotation_sheet(path: Path) -> str | None:
    sheets = workbook_sheets(path)
    if ANNOTATION_SHEET in sheets:
        return ANNOTATION_SHEET
    for sheet in sheets:
        try:
            df = pd.read_excel(path, sheet_name=sheet, nrows=1)
        except Exception:
            continue
        cols = {str(c).strip() for c in df.columns}
        if MIN_STANDARD_COLUMNS.issubset(cols):
            return sheet
    return None


def discover_annotation_workbooks() -> list[Path]:
    candidates = sorted(GEO_ROOT.rglob("*.xlsx"))
    return [path for path in candidates if not path.name.startswith("~$") and find_annotation_sheet(path)]


def clean_value(value: object) -> object:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value.lower() in {"nan", "none"}:
            return pd.NA
    return value


def read_annotation(path: Path) -> pd.DataFrame:
    sheet = find_annotation_sheet(path)
    if sheet is None:
        raise ValueError(f"No annotation-like sheet found in {path}")
    df = pd.read_excel(path, sheet_name=sheet)
    df = df.rename(columns={c: str(c).strip() for c in df.columns})
    for col in STANDARD_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA
    df = df[STANDARD_COLUMNS + [c for c in df.columns if c not in STANDARD_COLUMNS]]
    df = df.map(clean_value)
    df["annotation_file"] = str(path)
    df["annotation_sheet"] = sheet
    df["annotation_folder_gse"] = next((part for part in path.parts if re.fullmatch(r"GSE\d+", part)), pd.NA)
    folder_gse = df["annotation_folder_gse"].dropna().iloc[0] if df["annotation_folder_gse"].notna().any() else pd.NA
    if pd.notna(folder_gse) and str(folder_gse).startswith("GSE"):
        expression_ids, _, _ = expression_sample_ids(str(folder_gse))
        sample_ids = df["sample_id"].dropna().astype(str).tolist()
        if expression_ids and set(sample_ids) == set(expression_ids):
            df["source_series_id"] = df["series_id"]
            df["series_id"] = str(folder_gse)
            df["series_id_normalized_from_folder"] = "Yes"
        else:
            df["source_series_id"] = df["series_id"]
            df["series_id_normalized_from_folder"] = "No"
    else:
        df["source_series_id"] = df["series_id"]
        df["series_id_normalized_from_folder"] = "No"
    return df


def find_dataset_root(series_id: str) -> Path | None:
    direct = GEO_ROOT / series_id
    if direct.exists():
        return direct
    matches = list(GEO_ROOT.rglob(series_id))
    return matches[0] if matches else None


def read_series_matrix_sample_ids(path: Path) -> list[str]:
    opener = gzip.open if path.suffix == ".gz" else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            if line.startswith('"ID_REF"') or line.startswith("ID_REF"):
                return [part.strip().strip('"') for part in line.rstrip("\n").split("\t")[1:]]
    return []


def read_csv_header_sample_ids(path: Path) -> list[str]:
    opener = gzip.open if path.suffix == ".gz" else open
    with opener(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        header = next(csv.reader(handle))
    return [x.strip() for x in header[1:]]


def expression_sample_ids(series_id: str) -> tuple[list[str], str, str]:
    root = find_dataset_root(series_id)
    if root is None:
        return [], "", "dataset_root_not_found"

    # GSE150910's usable gene-level expression matrix is a supplementary count file.
    if series_id == "GSE150910":
        count_path = root / "supplementary" / "GSE150910_gene-level_count_file.csv.gz"
        if count_path.exists():
            return read_csv_header_sample_ids(count_path), str(count_path), "supplementary_gene_level_count"

    matrix_dir = root / "matrix"
    matrix_paths = []
    if matrix_dir.exists():
        matrix_paths = sorted(matrix_dir.glob(f"{series_id}*_series_matrix.txt.gz")) + sorted(
            matrix_dir.glob(f"{series_id}*_series_matrix.txt")
        )
    if not matrix_paths:
        matrix_paths = sorted(root.rglob(f"{series_id}*_series_matrix.txt.gz")) + sorted(
            root.rglob(f"{series_id}*_series_matrix.txt")
        )

    all_ids: list[str] = []
    used_paths = []
    for path in matrix_paths:
        ids = read_series_matrix_sample_ids(path)
        if ids:
            all_ids.extend(ids)
            used_paths.append(str(path))

    # Preserve order but remove duplicates.
    deduped = list(dict.fromkeys(all_ids))
    if deduped:
        return deduped, ";".join(used_paths), "geo_series_matrix"
    return [], ";".join(str(p) for p in matrix_paths), "expression_header_not_found"


def check_dataset(df: pd.DataFrame) -> dict[str, object]:
    series_id = str(df["series_id"].dropna().iloc[0]) if df["series_id"].notna().any() else "NA"
    sample_ids = df["sample_id"].dropna().astype(str).tolist()
    expr_ids, expr_file, expr_type = expression_sample_ids(series_id)
    missing_in_annotation = sorted(set(expr_ids) - set(sample_ids))
    missing_in_expression = sorted(set(sample_ids) - set(expr_ids))
    folder_gses = sorted(set(df["annotation_folder_gse"].dropna().astype(str)))
    organisms = sorted(set(df["organism"].dropna().astype(str)))
    include_values = sorted(set(df["include"].dropna().astype(str)))
    duplicate_sample_count = int(pd.Series(sample_ids).duplicated().sum())

    return {
        "series_id": series_id,
        "annotation_rows": int(len(df)),
        "unique_sample_ids": int(len(set(sample_ids))),
        "duplicate_sample_ids": duplicate_sample_count,
        "expression_sample_count": int(len(expr_ids)),
        "missing_in_annotation_count": int(len(missing_in_annotation)),
        "missing_in_expression_count": int(len(missing_in_expression)),
        "same_order": bool(expr_ids == sample_ids) if expr_ids else False,
        "match_status": "PASS" if expr_ids and not missing_in_annotation and not missing_in_expression else "FAIL",
        "expression_type": expr_type,
        "expression_file": expr_file,
        "annotation_folder_gse": ";".join(folder_gses) if folder_gses else "NA",
        "folder_matches_series_id": bool(series_id in folder_gses) if folder_gses else False,
        "annotation_sheets": ";".join(sorted(set(df["annotation_sheet"].dropna().astype(str)))),
        "source_series_ids": ";".join(sorted(set(df["source_series_id"].dropna().astype(str)))),
        "series_id_normalized_from_folder": ";".join(sorted(set(df["series_id_normalized_from_folder"].dropna().astype(str)))),
        "organism": ";".join(organisms) if organisms else "NA",
        "is_human": bool(organisms and set(organisms) == {"Homo sapiens"}),
        "include_values": ";".join(include_values) if include_values else "NA",
        "missing_in_annotation_examples": ";".join(missing_in_annotation[:10]),
        "missing_in_expression_examples": ";".join(missing_in_expression[:10]),
    }


def main() -> None:
    METADATA_ROOT.mkdir(parents=True, exist_ok=True)
    LOG_ROOT.mkdir(parents=True, exist_ok=True)

    paths = discover_annotation_workbooks()
    if not paths:
        raise SystemExit("No annotation_analysis sheets found.")

    frames = [read_annotation(path) for path in paths]
    merged = pd.concat(frames, ignore_index=True)
    merged["series_id"] = merged["series_id"].astype(str).str.strip()
    merged = merged[merged["series_id"].str.fullmatch(r"GSE\d+", na=False)].copy()

    qc_rows = []
    for series_id, sub in merged.groupby("series_id", sort=True):
        qc_rows.append(check_dataset(sub))
    qc = pd.DataFrame(qc_rows).sort_values("series_id")

    merged_out = METADATA_ROOT / "all_bulk_mirna_annotation.csv"
    qc_out = METADATA_ROOT / "all_bulk_mirna_annotation_qc.csv"
    xlsx_out = METADATA_ROOT / "all_bulk_mirna_annotation.xlsx"
    discovered_out = LOG_ROOT / "annotation_workbooks_discovered.json"

    merged.to_csv(merged_out, index=False, encoding="utf-8")
    qc.to_csv(qc_out, index=False, encoding="utf-8")
    with pd.ExcelWriter(xlsx_out, engine="openpyxl") as writer:
        merged.fillna("NA").replace("", "NA").to_excel(writer, sheet_name="annotation", index=False)
        qc.fillna("NA").replace("", "NA").to_excel(writer, sheet_name="qc", index=False)
    discovered_out.write_text(json.dumps([str(p) for p in paths], indent=2), encoding="utf-8")

    print(f"Wrote {merged_out} rows={len(merged)} datasets={merged['series_id'].nunique()}")
    print(f"Wrote {qc_out}")
    print(f"Wrote {xlsx_out}")
    print("\nQC summary:")
    print(
        qc[
            [
                "series_id",
                "match_status",
                "annotation_rows",
                "expression_sample_count",
                "missing_in_annotation_count",
                "missing_in_expression_count",
                "same_order",
                "is_human",
                "folder_matches_series_id",
            ]
        ].to_string(index=False)
    )


if __name__ == "__main__":
    main()
