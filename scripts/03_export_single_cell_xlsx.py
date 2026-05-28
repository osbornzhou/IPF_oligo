#!/usr/bin/env python
"""Export one single-cell annotation workbook per GSE.

Each workbook contains:
- sample_info
- cell_info

All missing or empty values are written as the literal string "NA".
"""

from __future__ import annotations

import csv
import gzip
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SC_ROOT = PROJECT_ROOT / "metadata" / "single_cell"
OUT_ROOT = SC_ROOT / "xlsx"

SAMPLE_PATH = SC_ROOT / "single_cell_sample_annotation.csv"
CELL_PATH = SC_ROOT / "single_cell_cell_annotation.csv.gz"

SAMPLE_COLUMNS = [
    "sample_id",
    "series_id",
    "geo_accession",
    "sample_title",
    "data_type",
    "dataset_role",
    "tissue",
    "organism",
    "group",
    "subgroup",
    "include",
    "platform",
    "library_technology",
    "library_id",
    "subject_id",
    "batch",
    "source",
    "n_cells",
    "cell_type_major_count",
    "cell_type_minor_count",
    "notes",
]

CELL_COLUMNS = [
    "cell_id",
    "series_id",
    "sample_id",
    "subject_id",
    "library_id",
    "geo_accession",
    "group",
    "subgroup",
    "include",
    "platform",
    "library_technology",
    "cell_type_major",
    "cell_type_minor",
    "cell_type_subclass",
    "nUMI",
    "nGene",
    "percent_mt",
    "batch",
    "source",
    "notes",
]


def clean_value(value: object) -> object:
    if pd.isna(value):
        return "NA"
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else "NA"
    return value


def clean_dataframe(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = "NA"
    out = out[columns]
    out = out.map(clean_value)
    return out


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, columns: list[str]) -> None:
    width_by_name = {
        "cell_id": 32,
        "sample_id": 18,
        "series_id": 14,
        "geo_accession": 16,
        "sample_title": 24,
        "cell_type_major": 20,
        "cell_type_minor": 30,
        "cell_type_subclass": 34,
        "notes": 60,
    }
    for idx, col in enumerate(columns, start=1):
        width = width_by_name.get(col, min(max(len(col) + 2, 12), 24))
        ws.column_dimensions[get_column_letter(idx)].width = width


def append_dataframe(ws, df: pd.DataFrame, columns: list[str]) -> None:
    ws.append(columns)
    for row in df.itertuples(index=False, name=None):
        ws.append([clean_value(x) for x in row])


def write_cell_rows(ws, series_id: str) -> int:
    ws.append(CELL_COLUMNS)
    count = 0
    found = False
    with gzip.open(CELL_PATH, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        for raw in reader:
            if raw.get("series_id") != series_id:
                continue
            found = True
            ws.append([clean_value(raw.get(col, "NA")) for col in CELL_COLUMNS])
            count += 1

    if not found:
        placeholder = {col: "NA" for col in CELL_COLUMNS}
        placeholder["series_id"] = series_id
        placeholder["notes"] = "cell-level metadata not available locally"
        ws.append([placeholder[col] for col in CELL_COLUMNS])
        count = 1
    return count


def assert_no_blank_cells(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    problems: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                    problems.append(f"{ws.title}!{cell.coordinate}")
                    if len(problems) >= 20:
                        return problems
    return problems


def main() -> None:
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    sample_all = pd.read_csv(SAMPLE_PATH)
    exported = []

    for series_id in sorted(sample_all["series_id"].unique()):
        sample_df = clean_dataframe(sample_all[sample_all["series_id"] == series_id], SAMPLE_COLUMNS)

        wb = Workbook(write_only=False)
        default = wb.active
        wb.remove(default)

        ws_sample = wb.create_sheet("sample_info")
        append_dataframe(ws_sample, sample_df, SAMPLE_COLUMNS)
        style_header(ws_sample)
        set_widths(ws_sample, SAMPLE_COLUMNS)

        ws_cell = wb.create_sheet("cell_info")
        n_cell_rows = write_cell_rows(ws_cell, series_id)
        style_header(ws_cell)
        set_widths(ws_cell, CELL_COLUMNS)

        out_path = OUT_ROOT / f"{series_id}_single_cell_annotation.xlsx"
        try:
            wb.save(out_path)
        except PermissionError:
            out_path = OUT_ROOT / f"{series_id}_single_cell_annotation_platform_updated.xlsx"
            wb.save(out_path)

        blanks = assert_no_blank_cells(out_path)
        if blanks:
            raise RuntimeError(f"Blank cells found in {out_path}: {blanks[:5]}")

        exported.append(
            {
                "series_id": series_id,
                "xlsx_path": str(out_path),
                "sample_rows": len(sample_df),
                "cell_rows": n_cell_rows,
                "blank_cells": 0,
            }
        )
        print(f"Wrote {out_path} sample_rows={len(sample_df)} cell_rows={n_cell_rows}")

    pd.DataFrame(exported).to_csv(OUT_ROOT / "single_cell_xlsx_manifest.csv", index=False, encoding="utf-8")


if __name__ == "__main__":
    main()
